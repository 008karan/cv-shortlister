#!/usr/bin/env python3
"""
compile_results.py — Merge Claude Code score JSONs into a ranked Excel report.

Reads:  output/batches/scores/cv_NNN_score.json  (written by Claude Code)
Writes: output/final/shortlist_YYYY-MM-DD.xlsx

Usage:
    python compile_results.py
    python compile_results.py --cutoff 25   # shortlist top 25% (default 30)
    python compile_results.py --top 15      # shortlist exactly top 15
"""

import argparse
import json
import math
import re
import sys
from datetime import date
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("✗  openpyxl not found. Install with: pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent
SCORES_DIR    = BASE_DIR / "output" / "batches" / "scores"
FINAL_DIR     = BASE_DIR / "output" / "final"
PROGRESS_PATH = BASE_DIR / "state" / "progress.json"

# ── Styles ─────────────────────────────────────────────────────────────────────
GREEN_FILL    = PatternFill("solid", fgColor="C6EFCE")
RED_FILL      = PatternFill("solid", fgColor="FFC7CE")
YELLOW_FILL   = PatternFill("solid", fgColor="FFEB9C")   # shortlisted + red flag
TEAL_FILL     = PatternFill("solid", fgColor="C6F4F0")   # research bonus candidate
HEADER_FILL   = PatternFill("solid", fgColor="2F4F7F")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
NORMAL_FONT = Font(size=10)
BOLD_FONT   = Font(bold=True, size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")
CTR_ALIGN   = Alignment(horizontal="center", vertical="top")
HDR_ALIGN   = Alignment(horizontal="center", vertical="center", wrap_text=True)

THIN = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)

# (header label, column width)
COLUMNS = [
    ("Rank",         6),
    ("Name",        22),
    ("File",        24),
    ("Total Score", 13),
    ("Must-Have",   13),
    ("Experience",  13),
    ("Nice-to-Have",14),
    ("Domain",      10),
    ("Shortlisted", 13),
    ("Research",    11),
    ("Red Flag",    11),
    ("Summary",     48),
]

# Columns that get center-aligned (1-based)
CENTER_COLS = {1, 4, 5, 6, 7, 8, 9, 10, 11}


# ── Helpers ────────────────────────────────────────────────────────────────────

def name_from_filename(filename: str) -> str:
    stem = Path(filename).stem.strip()
    stem = re.sub(r"[_\-]+", " ", stem)
    stem = re.sub(r"\s+", " ", stem).strip()
    return stem.title()


def load_score_files(scores_dir: Path) -> list[dict]:
    score_files = sorted(scores_dir.glob("*_score.json"))
    if not score_files:
        print(f"✗  No score files found in {scores_dir}")
        print("    Have Claude Code score the CVs first.")
        sys.exit(1)

    rows = []
    failed = 0
    for sf in score_files:
        try:
            with open(sf) as f:
                data = json.load(f)
            data["_score_file"] = sf.name
            rows.append(data)
        except Exception as e:
            print(f"  ⚠  Could not read {sf.name}: {e}")
            failed += 1

    print(f"  Loaded {len(rows)} score file(s)"
          + (f"  ({failed} unreadable)" if failed else ""))
    return rows


def build_rows(raw: list[dict]) -> tuple[list[dict], int]:
    rows = []
    parse_errors = 0

    for entry in raw:
        try:
            def _num(key, default=0):
                try:
                    return float(entry.get(key, default))
                except (TypeError, ValueError):
                    return float(default)

            must_have  = _num("must_have_match")
            experience = _num("experience_fit")
            nice       = _num("nice_to_have_match")
            domain     = _num("domain_relevance")

            # Recompute to guard against rounding drift
            weighted = round(
                must_have * 0.40 + experience * 0.25 + nice * 0.20 + domain * 0.15, 2
            )

            filename = entry.get("filename", entry.get("_score_file", "unknown"))

            rows.append({
                "filename":      filename,
                "name":          entry.get("name") or name_from_filename(filename),
                "must_have":     must_have,
                "experience":    experience,
                "nice":          nice,
                "domain":        domain,
                "weighted":      weighted,
                "research_flag": bool(entry.get("research_flag", False)),
                "red_flag":      bool(entry.get("red_flag", False)),
                "summary":       entry.get("one_line_summary", ""),
            })
        except Exception as e:
            print(f"  ⚠  Skipping malformed score entry: {e}")
            parse_errors += 1

    return rows, parse_errors


def apply_shortlist(rows: list[dict], cutoff_pct: int, top_n: Optional[int]) -> tuple[list[dict], int]:
    rows.sort(key=lambda r: r["weighted"], reverse=True)
    total = len(rows)
    threshold = top_n if top_n is not None else max(1, math.ceil(total * cutoff_pct / 100))

    for i, row in enumerate(rows):
        row["rank"] = i + 1
        row["shortlisted"] = "Yes" if (i + 1) <= threshold else "No"

    return rows, threshold


def write_sheet(ws, rows: list[dict], freeze: bool = True):
    ws.row_dimensions[1].height = 30
    for col_idx, (label, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HDR_ALIGN
        cell.border    = THIN
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if freeze:
        ws.freeze_panes = "A2"

    for row_idx, row in enumerate(rows, start=2):
        shortlisted   = row.get("shortlisted", "No")
        red_flag      = row.get("red_flag", False)
        research_flag = row.get("research_flag", False)

        if red_flag and shortlisted == "Yes":
            fill = YELLOW_FILL
        elif red_flag:
            fill = RED_FILL
        elif research_flag and shortlisted == "Yes":
            fill = TEAL_FILL     # research shortlisted — stand out visually
        elif shortlisted == "Yes":
            fill = GREEN_FILL
        else:
            fill = None

        ws.row_dimensions[row_idx].height = 30

        values = [
            row.get("rank", ""),
            row.get("name", ""),
            row.get("filename", ""),
            row.get("weighted", ""),
            row.get("must_have", ""),
            row.get("experience", ""),
            row.get("nice", ""),
            row.get("domain", ""),
            shortlisted,
            "Yes" if research_flag else "No",
            "Yes" if red_flag else "No",
            row.get("summary", ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN
            cell.font   = BOLD_FONT if col_idx == 4 else NORMAL_FONT
            if fill:
                cell.fill = fill
            cell.alignment = CTR_ALIGN if col_idx in CENTER_COLS else Alignment(wrap_text=True, vertical="top")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Compile Claude Code scores into a ranked Excel report")
    parser.add_argument("--cutoff", type=int, default=30, metavar="PCT",
                        help="Shortlist top N%% (default: 30)")
    parser.add_argument("--top", type=int, default=None, metavar="N",
                        help="Shortlist exactly top N candidates (overrides --cutoff)")
    args = parser.parse_args()

    print(f"\n{'='*56}")
    print("  compile_results.py — Build Shortlist Report")
    print(f"{'='*56}\n")

    print(f"📂  Reading scores from {SCORES_DIR.relative_to(BASE_DIR)}/")
    raw = load_score_files(SCORES_DIR)

    rows, parse_errors = build_rows(raw)
    rows, threshold    = apply_shortlist(rows, args.cutoff, args.top)

    shortlisted_count = sum(1 for r in rows if r["shortlisted"] == "Yes")
    red_flag_count    = sum(1 for r in rows if r["red_flag"])

    # Failed count from progress.json
    progress_failed = 0
    if PROGRESS_PATH.exists():
        with open(PROGRESS_PATH) as f:
            progress = json.load(f)
        progress_failed = len(progress.get("failed", []))

    # ── Build workbook ─────────────────────────────────────────────────────────
    FINAL_DIR.mkdir(parents=True, exist_ok=True)
    today    = date.today().strftime("%Y-%m-%d")
    out_path = FINAL_DIR / f"shortlist_{today}.xlsx"

    wb    = openpyxl.Workbook()
    ws_all = wb.active
    ws_all.title = "All Candidates"
    write_sheet(ws_all, rows)

    ws_rf = wb.create_sheet("Red Flags")
    rf_rows = [r for r in rows if r["red_flag"]]
    if rf_rows:
        write_sheet(ws_rf, rf_rows)
    else:
        ws_rf.cell(row=1, column=1, value="No red-flagged candidates.")

    wb.save(out_path)

    # ── Summary ────────────────────────────────────────────────────────────────
    cutoff_label = (
        f"top {args.top}" if args.top
        else f"top {args.cutoff}% → {threshold} of {len(rows)}"
    )

    print(f"\n{'='*56}")
    print(f"  Total evaluated : {len(rows)}")
    print(f"  Shortlisted     : {shortlisted_count}  ({cutoff_label})")
    print(f"  Red flags       : {red_flag_count}")
    print(f"  Parse errors    : {parse_errors + progress_failed}"
          f"  (see state/progress.json)")
    print(f"{'='*56}")
    print(f"\n✅  Saved → {out_path.relative_to(BASE_DIR)}\n")


if __name__ == "__main__":
    main()
